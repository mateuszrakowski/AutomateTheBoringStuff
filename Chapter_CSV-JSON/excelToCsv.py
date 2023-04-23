"""
Convert Excel files (including every sheet) to CSV files in current directory.

Usage:
python excelToCsv.py
"""

import openpyxl, csv, os


def findExcels():
    '''
    Searches for .xlsx files inside current working directory.

    Returns:
    list: List of excel filenames.
    '''
    excelFiles = []

    for filename in os.listdir("."):
        if filename.endswith(".xlsx"):
            excelFiles.append(filename)

    return excelFiles


def excelToCsv(listOfExcels):
    '''
    Creates CSV files from .xlsx files and its worksheets.

    Returns:
    str: Failure or success with list of converted .xlsx files.
    '''
    if listOfExcels == []:
        return "Couldn't find any .xlsx documents!"

    for excel in listOfExcels:
        wb = openpyxl.load_workbook(excel)
        for sheet in wb:
            csvFileObj = open(f"{excel}_{sheet}", "w", newline="")
            csvWriter = csv.writer(csvFileObj)

            for rowNum in range(1, sheet.max_row + 1):
                rowData = []
                for colNum in range(1, sheet.max_column + 1):
                    rowData.append(sheet.cell(row=rowNum, column=colNum).value)
                csvWriter.writerow(rowData)
            csvFileObj.close()

    return f"Excel files: {listOfExcels} converted to CSV files!"


if __name__ == "__main__":
    print(excelToCsv(findExcels()))
