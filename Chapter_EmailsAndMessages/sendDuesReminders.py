'''
Sends emails based on payment status in spreadsheet.

Usage:
python3 sendDuesReminders.py <email password>
'''

import openpyxl, smtplib, sys


# Open the spreadsheet and get the latest dues status
wb = openpyxl.load_workbook("duesRecords.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status
unpaidMembers = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != "paid":
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Log in to email account
smtpObj = smtplib.SMTP("smtp.server.com", "587")
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login("your@email.com", sys.argv[1])

# Send out reminder emails
for name, email in unpaidMembers.items():
    body = f"Subject: {latestMonth} dues unpaid.\nDear {name},\nRecords show that you have not paid dues for {latestMonth}. Please make this payment as soon as possible. Thank You!"
    print(f"Sending email to {email}...")
    sendmailStatus = smtpObj.sendmail("your@email.com", email, body)

    if sendmailStatus != {}:
        print(f"There was a problem sending email to {email}: {sendmailStatus}.")
smtpObj.quit()
