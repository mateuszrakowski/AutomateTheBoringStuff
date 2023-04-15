# spreadsheetToTxt.py - Write contents of provided spreadsheet to text file.

import openpyxl, sys
from openpyxl.utils.exceptions import InvalidFileException


# Check CLI arguments
if len(sys.argv) < 2:
    sys.exit("Wrong number of arguments!")

# Open workbook from provided file name
try:
    wb = openpyxl.load_workbook(sys.argv[1])
    sheet = wb.active
except InvalidFileException:
    sys.exit("Invalid file format.")
except FileNotFoundError:
    sys.exit("File doesn't exist.")

# Create list to store cells values
columnsList = []

# Read spreadsheet cells
for colNum in range(1, sheet.max_column + 1):
    # Create nested lists - one list represents one column values
    columnsList.append([sheet.cell(row=rowNum, column=colNum).value for rowNum in range(1, sheet.max_row + 1)])

# For every column create one file
for fileNumber in range(len(columnsList)):
    with open(f"file{fileNumber}.txt", "w") as file:
        # Write one line per one list element
        for fileLine in columnsList[fileNumber]:
            if fileLine is not None:
                file.write(fileLine + "\n")
