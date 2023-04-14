# multiplicationTable.py - Creates a 6x6 table with multiplication of numbers

import openpyxl, sys
from openpyxl.styles import Font


# Check valid argument number
if len(sys.argv) == 2:
    # Adding 2 to argument is necessary as we starting (skipping two iterations) loop from second column & row
    tableRange = int(sys.argv[1]) + 2
else:
    sys.exit("Wrong number of arguments.")

# Create workbook and set active sheet
wb = openpyxl.Workbook()
sheet = wb.active

# Change sheet title
sheet.title = "Multiplication table"

# Create font variable
fontBold = Font(bold=True)

# These two separate loops could be merged into one but this way is clearer in my opinion
# Create labels starting from column "B" and row 2
for columnNum in range(2, tableRange):
    sheet.cell(row=1, column=columnNum).value = columnNum - 1
    sheet.cell(row=columnNum, column=1).value = columnNum - 1

for rowNum in range(2, tableRange):
    for columnNum in range(2, tableRange):
        # Save label cells to variable for better transparency
        mainRow = sheet.cell(row=1, column=columnNum)
        mainColumn = sheet.cell(row=rowNum, column=1)

        # Set label font to bold
        mainRow.font, mainColumn.font = fontBold, fontBold

        # Calculate multiplication
        sheet.cell(row=rowNum, column=columnNum).value = mainRow.value * mainColumn.value

wb.save("Multiplication Table.xlsx")
