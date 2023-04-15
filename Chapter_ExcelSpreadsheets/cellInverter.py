# cellInverter.py - Inverts the row and column of the cells in the spreadsheet.

import openpyxl, sys
from openpyxl.utils.exceptions import InvalidFileException


# Check CLI arguments
if len(sys.argv) != 2:
    sys.exit("Invalid number of arguments.")

# Open spreadsheet
try:
    wb = openpyxl.load_workbook(sys.argv[1])
    sheet = wb.active
except InvalidFileException:
    sys.exit("Invalid file format.")
except FileNotFoundError:
    sys.exit("File doesn't exist.")

# Create list to store cells values
cellsList = []

# Create new spreadsheet
wbNew = openpyxl.Workbook()
sheetNew = wbNew.active

# Change title
sheetNew.title = sheet.title
    
# Read cells from original spreadsheet
for colNum in range(1, sheet.max_column + 1):
    # Create nested lists - one list represents one column
    cellsList.append([sheet.cell(row=rowNum, column=colNum).value for rowNum in range(1, sheet.max_row + 1)])

    # Write inverted cells to new spreadsheet
    for rowNum in range(1, sheet.max_row + 1):
        sheetNew.cell(row=colNum, column=rowNum).value = cellsList[colNum - 1][rowNum - 1]

wbNew.save("New_" + sys.argv[1])
