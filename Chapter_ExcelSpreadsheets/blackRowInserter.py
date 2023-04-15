# blackRowInserter.py - In a given row number (first argument) insert provided number of blank rows (second argument)

import openpyxl, sys
from openpyxl.utils.exceptions import InvalidFileException


# Check CLI arguments
if len(sys.argv) == 4:
    rowNumber = int(sys.argv[1])
    blanksNumber = int(sys.argv[2])
else:
    sys.exit("Wrong number of arguments!")

# Open workbook from provided file name
try:
    wb = openpyxl.load_workbook(sys.argv[3])
except InvalidFileException:
    sys.exit("Invalid file format.")
except FileNotFoundError:
    sys.exit("File doesn't exist.")

sheet = wb.active
title = sheet.title

# Create new spreadsheet
wbNew = openpyxl.Workbook()
sheetNew = wbNew.active

# Copy and set title from original sheet
sheetNew.title = title

# The following solution to the problem is created intentionally to not use provided insert method from openpyxl modules. This gave me a better understanting of performing actions on sheets and cells. Second version of this program will cover much easier approach by using the mentioned function.

# Read every cell in row
for rowNum in range(1, sheet.max_row + blanksNumber):
    # Read every cell in column
    for columnNum in range(1, sheet.max_column + 1):
        # Save present cell from original sheet to variable
        presentCell = sheet.cell(row=rowNum, column=columnNum).value
            
        # For provided number of blank rows
        if rowNum in range(rowNumber, rowNumber + blanksNumber):
            # Move (copy) cells before making them blank to further (row number + blank rows) number of rows
            sheetNew.cell(row=rowNum + blanksNumber, column=columnNum).value = presentCell
            # Insert blank rows by creating empty cell by cell
            sheetNew.cell(row=rowNum, column=columnNum).value = ""
        
        # After inserting the blank rows continue to copy rest of cells
        elif rowNum >= rowNumber + blanksNumber:
            sheetNew.cell(row=rowNum + blanksNumber, column=columnNum).value = presentCell

        # Copy cells before insert row
        else:
            sheetNew.cell(row=rowNum, column=columnNum).value = presentCell

wbNew.save("New_" + sys.argv[3])
